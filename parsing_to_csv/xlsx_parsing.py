from openpyxl import load_workbook
import csv

# an example of reading data from an .xlsx file with Python, using the "openpyxl" library.
# the source data can be composed and downloaded from:
# https://fred.stlouisfed.org/series/U6RATE

#Firt: pass the filename as an ingredient to the load_workbook recipe
source_workbook = load_workbook(filename='fredgraph.xlsx')
#print sheetname
print(source_workbook.sheetnames)

#loop throw worksheet

for sheet_num, sheet_name in enumerate(source_workbook.sheetnames):
    #variable that points to current worksheet
    current_sheet = source_workbook[sheet_name]
    print(sheet_name)

    #create output file
    output_file = open("xls"+sheet_name+".csv","w")

    #using the csv.writer to write rows of data to the output_file
    output_writer = csv.writer(output_file)

    #loop through every row in the sheet:
    for row in current_sheet.iter_rows():
        #create a empty list, to put the actual values of the cells in each row
        empty_cells = []

        #for every cell (or column) in each row
        for cell in row:
            print(cell, cell.value)
            #add the cell value to the list
            empty_cells.append(cell.value)

        #add (write) the new constructed data row to the output file
        output_writer.writerow(empty_cells)

    #close the .csv file
    output_file.close()




